from openpyxl import Workbook
from .models import CustomerLegalEntity


def get_data():
    legal_entities = CustomerLegalEntity.objects.all()

    data = []

    for entity in legal_entities:
        ceo = entity.ceo


        row_data = (
            ceo.fullname if ceo else '',  # ФИО руководителя
            ceo.email if ceo else '',  # Электронная почта руководителя
            entity.inn,  # ИНН юридического лица
            entity.okpo,  # Код ОКПО
            entity.registration_number,  # Регистрационный номер
            entity.fax,  # Факс
        )

        data.append(row_data)

    return data


class ExcelPropertyGenerator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None

    def load_workbook(self):
        self.workbook = Workbook()
        self.ws = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)
        return self.file_path

    def generate_report(self, data):
        self.load_workbook()

        # Заголовки столбцов
        headers = [
            'ФИО Руководителя',
            'Электронная почта Руководителя',
            'ИНН',
            'Код ОКПО',
            'Регистрационный номер',
            'Факс',
        ]

        self.ws.append(headers)

        for row_data in data:
            self.ws.append(row_data)

        self.save_workbook()



if __name__ == "__main__":
    excel_file_path = "Excel.xlsx"
    data = get_data()

    excel_generator = ExcelPropertyGenerator(excel_file_path)
    excel_generator.generate_report(data)