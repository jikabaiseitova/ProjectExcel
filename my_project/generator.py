from openpyxl import load_workbook
from .models import CustomerLegalEntity


class ExcelPropertyGenerator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None

    @staticmethod
    def get_data():
        data = []
        legal_entities = CustomerLegalEntity.objects.all()

        for entity in legal_entities:
            data.append(
                (entity.fullname, entity.email, entity.inn, entity.okpo, entity.registration_number, entity.fax))

        return data

    def load_workbook(self):
        self.workbook = load_workbook(self.file_path)  ## Открытие книги, активировать
        self.ws = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)    ### Сохранение файла, переписывает файл если есть такой, если нету - создает
        return self.file_path

    def generator(self):              ### Логика, в дате хранится результат функции, список из множеств tuple
        data = self.get_data()
        for row_data in data:
            self.ws.append(row_data)
            print(row_data)

    def generate_report(self):        ## Метод - вызывает функцию
        self.load_workbook()    ## открытие киги
        self.generator()        ## генерирует
        self.save_workbook()    ## закрытие книги
