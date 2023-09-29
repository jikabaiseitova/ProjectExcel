from openpyxl import load_workbook
from django.http import HttpResponse
from models import *


class ExcelPropertyGenerator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None

    def load_workbook(self):
        self.workbook = load_workbook(self.file_path)
        self.ws = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)
        return self.file_path

    def generate_report(self):
        self.load_workbook()
        data = get_data()

        for row_data in data:
            self.ws.append(row_data)

        self.save_workbook()


def get_data():
    data = []
    legal_entities = CustomerLegalEntity.objects.all()

    for entity in legal_entities:
        data.append((entity.fullname, entity.email, entity.inn, entity.okpo, entity.registration_number, entity.fax))

    return data


def generate_excel_report(request):
    file_path = "my_book.xlsx"
    generated_excel = ExcelPropertyGenerator(file_path)
    generated_excel.generate_report()

    with open(file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f"attachment; filename=отчет.xlsx"

    return response

