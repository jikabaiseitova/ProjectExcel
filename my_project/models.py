from django.db import models
from openpyxl import Workbook
from openpyxl import load_workbook

from my_project.write_excel import get_data


class Customer(models.Model):
    email = models.EmailField(
        verbose_name='Электронная почта',
        null=True, blank=True
    )
    fullname = models.CharField(
        max_length=100,
        verbose_name='ФИО клиента/Наименование компании',
        null=True, blank=True
    )

    def __str__(self):
        return self.fullname


class CustomerIndividual(Customer):
    last_name = models.CharField(
        max_length=100,
        verbose_name='Фамилия',
    )
    first_name = models.CharField(
        max_length=100,
        verbose_name='Имя',
    )
    patronymic = models.CharField(
        max_length=100,
        verbose_name='Отчество',
        null=True, blank=True
    )

    class Meta:
        verbose_name = 'Страхователь Физ.Лицо'
        verbose_name_plural = 'Страхователи Физ.Лица'

    def save(self, *args, **kwargs):
        if self.last_name:
            if self.patronymic:
                self.fullname = f'{self.last_name} {self.first_name} {self.patronymic}'
            else:
                self.fullname = f'{self.last_name} {self.first_name}'
        super().save(*args, **kwargs)


class CustomerLegalEntity(Customer):
    inn = models.CharField(
        max_length=25,
        verbose_name='ИНН',
    )
    okpo = models.CharField(
        max_length=25,
        verbose_name='Код ОКПО',
        help_text='Общегосударственный классификатор предприятий и организаций',
        null=True, blank=True
    )
    registration_number = models.CharField(
        max_length=255,
        verbose_name='Регистрационный номер',
        null=True, blank=True
    )
    fax = models.CharField(
        max_length=255,
        verbose_name='Факс',
        null=True, blank=True
    )
    ceo = models.ForeignKey(
        CustomerIndividual,
        on_delete=models.RESTRICT,
        verbose_name='Руководитель',
        help_text='Руководитель компании',
        related_name='legal_boss'
    )


class ExcelPropertyGenerator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None

    def load_workbook(self):
        self.workbook = load_workbook(self.file_path)
        self.ws = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)
        return self.file_path

    def generate_report(self):
        self.load_workbook()
        data = get_data()

        for row_data in data:
            self.ws.append(row_data)

        self.save_workbook()
