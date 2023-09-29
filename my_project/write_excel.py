from openpyxl import load_workbook

from .models import CustomerLegalEntity


def get_data():
    data = []
    legal_entities = CustomerLegalEntity.objects.all()

    for entity in legal_entities:
        data.append((entity.fullname, entity.email, entity.inn, entity.okpo, entity.registration_number, entity.fax))

    return data


class ExcelPropertyGenerator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None

    def load_workbook(self):
        self.workbook = load_workbook(self.file_path)
        self.ws = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)
        return self.file_path

    def generate_report(self):
        self.load_workbook()
        data = get_data()

        for row_data in data:
            self.ws.append(row_data)

        self.save_workbook()



